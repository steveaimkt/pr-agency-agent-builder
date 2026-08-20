#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
산출물의 표를 엑셀로 만든다.

  python3 tools/md2xlsx.py output/01-지표원장.md
  python3 tools/md2xlsx.py output/                 폴더 안 md 전부

같은 자리에 .xlsx 가 생긴다. 정렬하고 걸러 볼 수 있다.
색과 글꼴은 tools/design.md 를 따른다.

엑셀 뼈대로 짠다 · outputs.md 「엑셀로 갈 때」 와 같다
  시트 1  한눈에    숫자 카드. 이 파일이 무엇인지 한 화면에
  시트 2  원장      ⭐ 제일 긴 표. 한 줄이 한 건
  시트 3  집계      나머지 표를 위아래로 쌓는다
  시트 4  기준      판단 기준·목표치가 있으면

⛔ 표마다 시트를 만들지 않는다. 시트가 열 개를 넘으면 찾지를 못한다.

무엇이 되나
  머리행 고정 · 자동 필터 · 숫자는 오른쪽 정렬 · 열 너비 자동
  [FACT-CHECK] 와 [확인 필요] 는 형광으로 칠한다
"""
import sys, os, re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print('openpyxl 이 없습니다 · pip3 install openpyxl'); sys.exit(1)

# ─── 색 · tools/design.md 와 같은 값 ───
INK, BODY, META = '2A2A28', '4C4D59', '848383'
SOFT, LINE, HL = 'F0EDE8', 'E6E2DC', 'EBFF2C'
GOOD, BAD = '37624A', 'A83226'
FONT = 'Pretendard'

MARKS = ('[FACT-CHECK]', '[확인 필요]')
KPI_RE = re.compile(r'^([\d,]+(?:\.\d+)?[%가-힣]*)\s+(.+?)(?:\s*·\s*(좋음|나쁨))?\s*$')
NUM_RE = re.compile(r'^-?[\d,]+(?:\.\d+)?$')

thin = Side(style='thin', color=LINE)
BORDER_H = Border(bottom=thin)


def clean(s):
    return re.sub(r'\*\*|`', '', s).strip()


def sheet_name(raw, used):
    """엑셀 시트 이름 규칙 · 31자 · 아래 글자 금지"""
    s = re.sub(r'[\\/*?:\[\]]', ' ', clean(raw)).strip() or '표'
    s = s[:28]
    base, k = s, 2
    while s in used:
        s = f'{base[:26]} {k}'; k += 1
    used.add(s)
    return s


def parse(md):
    """제목을 따라가며 표와 숫자 블록을 뽑는다"""
    lines = md.split('\n')
    tables, kpis = [], []
    title = ''
    i, n = 0, len(lines)
    while i < n:
        ln = lines[i]
        m = re.match(r'^(#{1,4})\s+(.+)$', ln)
        if m:
            title = clean(m.group(2)); i += 1; continue

        m = re.match(r'^```(숫자|막대|달성)\s*(.*)$', ln.strip())
        if m:
            kind, t = m.group(1), clean(m.group(2))
            body = []
            i += 1
            while i < n and not lines[i].strip().startswith('```'):
                body.append(lines[i]); i += 1
            i += 1
            if kind == '숫자':
                kpis.append((t or title, body))
            else:
                rows = [['항목', '값']]
                for b in body:
                    mm = re.match(r'^(.+?)\s+([\d,]+(?:\.\d+)?)\s*\S*$', b.strip())
                    if mm:
                        rows.append([mm.group(1).strip(), mm.group(2)])
                if len(rows) > 2:
                    tables.append((t or title, rows))
            continue

        if ln.strip().startswith('```'):
            i += 1
            while i < n and not lines[i].strip().startswith('```'):
                i += 1
            i += 1
            continue

        if ln.strip().startswith('|') and i + 1 < n and re.match(r'^\s*\|[\s:|-]+\|\s*$', lines[i + 1]):
            rows = []
            while i < n and lines[i].strip().startswith('|'):
                cells = [c.strip() for c in lines[i].strip().strip('|').split('|')]
                if not re.match(r'^[\s:|-]+$', ''.join(cells)):
                    rows.append(cells)
                i += 1
            if len(rows) >= 2:
                tables.append((title, rows))
            continue
        i += 1
    return kpis, tables


def write_table(ws, rows):
    ws.sheet_view.showGridLines = False
    head = [clean(c) for c in rows[0]]
    ws.append(head)
    for c in range(1, len(head) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=FONT, size=9, bold=True, color=META)
        cell.fill = PatternFill('solid', fgColor=SOFT)
        cell.border = BORDER_H
        cell.alignment = Alignment(vertical='center', wrap_text=True)

    for r in rows[1:]:
        vals, marked = [], []
        for j in range(len(head)):
            raw = r[j] if j < len(r) else ''
            marked.append(any(k in raw for k in MARKS))
            v = clean(raw)
            if NUM_RE.match(v):
                try:
                    v = float(v.replace(',', '')) if '.' in v else int(v.replace(',', ''))
                except ValueError:
                    pass
            vals.append(v)
        ws.append(vals)
        rr = ws.max_row
        for c in range(1, len(head) + 1):
            cell = ws.cell(row=rr, column=c)
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.border = BORDER_H
            cell.alignment = Alignment(
                horizontal='right' if isinstance(cell.value, (int, float)) else 'left',
                vertical='center', wrap_text=True)
            if marked[c - 1]:
                cell.fill = PatternFill('solid', fgColor=HL)
                cell.font = Font(name=FONT, size=10, bold=True, color=INK)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(head))}{ws.max_row}'
    for c in range(1, len(head) + 1):
        longest = max((len(str(ws.cell(row=r, column=c).value or '')) for r in range(1, ws.max_row + 1)), default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(46, max(10, longest * 1.7))
    ws.row_dimensions[1].height = 26


def write_kpi(ws, kpis, src_name='', guide=None, notes=None):
    """한눈에 시트 · 숫자를 카드로 앉힌다. 표로 만들지 않는다."""
    ws.sheet_view.showGridLines = False
    edge = Side(style='thin', color=LINE)
    fill = PatternFill('solid', fgColor=SOFT)

    # 카드 한 장이 세 열을 쓰고, 사이에 한 열을 비운다
    CARD_W, GAP = 3, 1
    ws.column_dimensions['A'].width = 2
    r = 2
    for title, body in kpis:
        cards = []
        for ln in body:
            m = KPI_RE.match(ln.strip())
            if m:
                cards.append((m.group(1), m.group(2), m.group(3)))
        if not cards:
            continue

        ws.cell(row=r, column=2, value=clean(title) or '한눈에').font = Font(
            name=FONT, size=13, bold=True, color=INK)
        r += 2

        col = 2
        for big, label, judge in cards:
            c2, c3 = get_column_letter(col), get_column_letter(col + CARD_W - 1)
            judge_col = GOOD if judge == '좋음' else BAD if judge == '나쁨' else INK

            # 큰 숫자
            ws.merge_cells(f'{c2}{r}:{c3}{r}')
            cell = ws.cell(row=r, column=col, value=big)
            cell.font = Font(name=FONT, size=26, bold=True, color=judge_col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 라벨
            ws.merge_cells(f'{c2}{r+1}:{c3}{r+1}')
            lab = ws.cell(row=r + 1, column=col, value=clean(label))
            lab.font = Font(name=FONT, size=10, color=BODY)
            lab.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # 판정 한 줄
            ws.merge_cells(f'{c2}{r+2}:{c3}{r+2}')
            jg = ws.cell(row=r + 2, column=col, value=(judge or ''))
            jg.font = Font(name=FONT, size=9, bold=True, color=judge_col)
            jg.alignment = Alignment(horizontal='center', vertical='center')

            # 카드 배경 · 테두리는 바깥에만 두른다. 안쪽에 줄을 그으면 표로 보인다
            for rr in range(r, r + 3):
                for cc in range(col, col + CARD_W):
                    x = ws.cell(row=rr, column=cc)
                    x.fill = fill
                    x.border = Border(
                        top=edge if rr == r else None,
                        bottom=edge if rr == r + 2 else None,
                        left=edge if cc == col else None,
                        right=edge if cc == col + CARD_W - 1 else None)
                    ws.column_dimensions[get_column_letter(cc)].width = 9
            col += CARD_W + GAP
            ws.column_dimensions[get_column_letter(col - GAP)].width = 2

        ws.row_dimensions[r].height = 42
        ws.row_dimensions[r + 1].height = 20
        ws.row_dimensions[r + 2].height = 20
        r += 5

    # ── 이 파일에 무엇이 있나 ──
    if guide:
        ws.cell(row=r, column=2, value='이 파일에 무엇이 있나').font = Font(
            name=FONT, size=12, bold=True, color=INK)
        r += 2
        for j, h in enumerate(['시트', '무엇이 들어 있나', '건수'], start=2):
            c = ws.cell(row=r, column=j, value=h)
            c.font = Font(name=FONT, size=9, bold=True, color=META)
            c.fill = PatternFill('solid', fgColor=SOFT)
            c.border = BORDER_H
        r += 1
        for name, what, cnt in guide:
            ws.cell(row=r, column=2, value=name).font = Font(name=FONT, size=10, bold=True, color=INK)
            ws.cell(row=r, column=3, value=what).font = Font(name=FONT, size=10, color=BODY)
            c = ws.cell(row=r, column=4, value=cnt)
            c.font = Font(name=FONT, size=10, color=INK)
            c.alignment = Alignment(horizontal='right')
            for j in range(2, 5):
                ws.cell(row=r, column=j).border = BORDER_H
            r += 1
        r += 2

    # ── 눈에 띄는 것 ──
    if notes:
        ws.cell(row=r, column=2, value='눈에 띄는 것').font = Font(
            name=FONT, size=12, bold=True, color=INK)
        r += 2
        for txt, kind in notes:
            col = BAD if kind == '나쁨' else GOOD if kind == '좋음' else BODY
            c = ws.cell(row=r, column=2, value=('•  ' + txt))
            c.font = Font(name=FONT, size=10.5, color=col, bold=(kind == '나쁨'))
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            ws.row_dimensions[r].height = 20
            r += 1
        r += 2

    if src_name:
        c = ws.cell(row=r, column=2, value=f'원본 · {src_name}')
        c.font = Font(name=FONT, size=9, color=META)


def stack(ws, tables):
    """표 여럿을 한 시트에 위아래로 쌓는다. 제목을 앞에 둔다."""
    ws.sheet_view.showGridLines = False
    r = 1
    widest = 0
    for title, rows in tables:
        c0 = ws.cell(row=r, column=1, value=clean(title) or '표')
        c0.font = Font(name=FONT, size=12, bold=True, color=INK)
        r += 2
        head = [clean(c) for c in rows[0]]
        widest = max(widest, len(head))
        for j, h in enumerate(head, start=1):
            cell = ws.cell(row=r, column=j, value=h)
            cell.font = Font(name=FONT, size=9, bold=True, color=META)
            cell.fill = PatternFill('solid', fgColor=SOFT)
            cell.border = BORDER_H
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        r += 1
        for body in rows[1:]:
            for j in range(len(head)):
                raw = body[j] if j < len(body) else ''
                v = clean(raw)
                if NUM_RE.match(v):
                    try:
                        v = float(v.replace(',', '')) if '.' in v else int(v.replace(',', ''))
                    except ValueError:
                        pass
                cell = ws.cell(row=r, column=j + 1, value=v)
                cell.font = Font(name=FONT, size=10, color=INK)
                cell.border = BORDER_H
                cell.alignment = Alignment(
                    horizontal='right' if isinstance(v, (int, float)) else 'left',
                    vertical='center', wrap_text=True)
                if any(k in raw for k in MARKS):
                    cell.fill = PatternFill('solid', fgColor=HL)
                    cell.font = Font(name=FONT, size=10, bold=True, color=INK)
            r += 1
        r += 2
    for c in range(1, max(1, widest) + 1):
        longest = max((len(str(ws.cell(row=x, column=c).value or '')) for x in range(1, ws.max_row + 1)), default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(46, max(10, longest * 1.6))


def _notes(kpis, tables):
    """눈에 띄는 것을 뽑는다. 나쁨 판정과 확인 필요, 미달 달성률."""
    out = []
    for _, body in kpis:
        for ln in body:
            m = KPI_RE.match(ln.strip())
            if m and m.group(3):
                out.append((f'{m.group(2)} · {m.group(1)}', m.group(3)))
    marked = 0
    for _, rows in tables:
        for r in rows[1:]:
            marked += sum(1 for c in r if any(k in c for k in MARKS))
    if marked:
        out.append((f'확인이 필요한 칸이 {marked}개 있습니다. 형광으로 칠해 두었습니다.', '나쁨'))
    return out[:8]


def convert(src, dst):
    kpis, tables = parse(open(src, encoding='utf-8').read())
    if not kpis and not tables:
        return False, ''
    wb = Workbook(); wb.remove(wb.active)
    note, guide = [], []

    if tables:
        # 제일 긴 표가 원장이다. 한 줄이 한 건인 자리
        tables = sorted(tables, key=lambda x: -len(x[1]))
        led_title, led_rows = tables[0]
        # 원장은 「한 줄이 한 건」인 표다. 여덟 줄 넘고 세 열 넘어야 원장으로 본다.
        # 아니면 원장 시트를 만들지 않는다. 짧은 표를 원장이라 부르면 거짓말이 된다.
        if len(led_rows) >= 9 and len(led_rows[0]) >= 3:
            write_table(wb.create_sheet('원장'), led_rows)
            note.append(f'원장 {len(led_rows) - 1}건')
            guide.append(('원장', clean(led_title) or '한 줄이 한 건', f'{len(led_rows) - 1}건'))
            rest = tables[1:]
        else:
            rest = tables
        crit = [x for x in rest if re.search(r'기준|목표|합의|규정', clean(x[0]))]
        agg = [x for x in rest if x not in crit]
        if agg:
            stack(wb.create_sheet('집계'), agg)
            note.append(f'집계 {len(agg)}표')
            guide.append(('집계', ' · '.join(clean(a[0]) or '표' for a in agg[:3])[:40], f'{len(agg)}표'))
        if crit:
            stack(wb.create_sheet('기준'), crit)
            note.append(f'기준 {len(crit)}표')
            guide.append(('기준', ' · '.join(clean(c[0]) or '표' for c in crit[:3])[:40], f'{len(crit)}표'))

    if kpis:
        ws = wb.create_sheet('한눈에')
        write_kpi(ws, kpis, os.path.basename(src), guide, _notes(kpis, tables))
        wb.move_sheet('한눈에', offset=-(len(wb.sheetnames) - 1))
        note.insert(0, '한눈에')

    if not wb.sheetnames:
        return False, ''
    wb.save(dst)
    return True, ' · '.join(note)


def main():
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    target = sys.argv[1]
    files = []
    if os.path.isdir(target):
        files = [os.path.join(target, f) for f in sorted(os.listdir(target)) if f.endswith('.md')]
    elif target.endswith('.md'):
        files = [target]
    if not files:
        print(f'md 파일을 못 찾았습니다: {target}'); sys.exit(1)
    made = 0
    for f in files:
        out = f[:-3] + '.xlsx'
        ok, note = convert(f, out)
        if ok:
            print(f'  {os.path.basename(out):<28} {note}')
            made += 1
        else:
            print(f'  (건너뜀) {os.path.basename(f)} · 표가 없습니다')
    print(f'\n{made}개 만들었습니다. 정렬과 거르기가 됩니다.')


if __name__ == '__main__':
    main()
